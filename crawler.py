"""
GIGW 3.0 Web Crawler
====================
Crawls a website domain exhaustively (no depth limit) and records all
internal webpages. Uses a pandas DataFrame for bookkeeping and exports
results to a CSV (default) or a formatted Excel workbook.

The default CSV output (URL, Page Title, HTTP Status, …) is directly
compatible with the WCAG 1.1.1 alt-text checker (alt_text.py).

Dependencies:
    pip install requests beautifulsoup4 pandas openpyxl

Usage:
    python crawler.py
    python crawler.py --url https://example.gov.in
    python crawler.py --url https://example.gov.in --delay 1.5
    python crawler.py --url https://example.gov.in --output report.csv
    python crawler.py --url https://example.gov.in --format xlsx --output report.xlsx

Pipeline (crawl → WCAG check):
    python crawler.py --url https://example.gov.in --output crawl.csv
    python alt_text.py crawl.csv wcag_report.csv
"""

import requests
import argparse
import sys
import time
from urllib.parse import urljoin, urlparse, urldefrag
from collections import deque
from datetime import datetime

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("[ERROR] Run: pip install requests beautifulsoup4 pandas openpyxl")
    sys.exit(1)

# ── Constants ──────────────────────────────────────────────────────────────────

DEFAULT_DELAY      = 1.0
DEFAULT_TIMEOUT    = 15
DEFAULT_USER_AGENT = "GIGW-Crawler/2.0 (Government Website Evaluation Tool)"

LINK_TAGS = {"a": "href", "link": "href", "area": "href"}

SKIP_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".zip", ".rar", ".tar", ".gz",
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg", ".webp", ".ico",
    ".mp3", ".mp4", ".avi", ".mov", ".wmv",
    ".css", ".js", ".json", ".xml",
    ".woff", ".woff2", ".ttf", ".eot",
}

# ── Colour palette ─────────────────────────────────────────────────────────────
C_TITLE_BG  = "0D2137"
C_HEADER_BG = "1A3C5E"
C_HEADER_FG = "FFFFFF"
C_OK_BG     = "EAF4EC"
C_ALT_BG    = "F7FAFD"
C_ERR_BG    = "FCE8E8"
C_WARN_BG   = "FFF8E1"
C_META_BG   = "F0F4F8"
C_BORDER    = "C5D5E4"

# ── URL Utilities ──────────────────────────────────────────────────────────────

def normalize(url: str) -> str:
    url, _ = urldefrag(url)
    return url.rstrip("/")

def same_domain(url: str, base: str) -> bool:
    netloc = urlparse(url).netloc.lower()
    return netloc == base or netloc.endswith("." + base)

def is_webpage(url: str) -> bool:
    path = urlparse(url).path.lower()
    last = path.rsplit("/", 1)[-1]
    ext  = "." + last.rsplit(".", 1)[-1] if "." in last else ""
    return ext not in SKIP_EXTENSIONS

# ── Crawler ────────────────────────────────────────────────────────────────────

def crawl(start_url: str, delay: float = DEFAULT_DELAY,
          timeout: int = DEFAULT_TIMEOUT,
          max_depth: int = -1) -> pd.DataFrame:

    base    = urlparse(start_url).netloc.lower()
    session = requests.Session()
    session.headers.update({
        "User-Agent": DEFAULT_USER_AGENT,
        "Accept":     "text/html,*/*;q=0.8",
    })

    # Pandas DataFrame is the single source of truth
    columns = ["#", "URL", "Page Title", "HTTP Status",
               "Found On (Parent URL)", "Depth", "Discovered At", "Notes"]
    df = pd.DataFrame(columns=columns)

    queue  = deque()   # (url, parent_url, depth)
    queued = set()     # prevents duplicate queue entries

    def enqueue(url, parent, depth):
        nurl = normalize(url)
        if nurl not in queued and nurl not in df["URL"].values:
            queued.add(nurl)
            queue.append((nurl, parent, depth))

    enqueue(normalize(start_url), "—", 0)

    depth_label = "Unlimited (exhaustive)" if max_depth == -1 else str(max_depth)
    print(f"\n{'='*65}")
    print(f"  GIGW 3.0 Web Crawler")
    print(f"{'='*65}")
    print(f"  Start URL : {start_url}")
    print(f"  Domain    : {base}")
    print(f"  Max Depth : {depth_label}")
    print(f"{'='*65}\n")

    while queue:
        url, parent, depth = queue.popleft()

        # Guard against race between queue and df
        if url in df["URL"].values:
            continue

        row_num = len(df) + 1
        ts      = datetime.now().strftime("%H:%M:%S")
        print(f"  [{row_num:>4}] (depth={depth}) {url}")

        # ── Fetch ──────────────────────────────────────────────────────────────
        try:
            resp   = session.get(url, timeout=timeout, allow_redirects=True)
            final  = normalize(resp.url)
            status = resp.status_code
            ctype  = resp.headers.get("Content-Type", "")
            notes  = ""

            if not same_domain(final, base):
                notes = "Redirected to external domain"
                df.loc[len(df)] = [row_num, url, "(external redirect)",
                                   status, parent, depth, ts, notes]
                time.sleep(delay)
                continue

            if status != 200:
                notes = f"HTTP {status}"
                df.loc[len(df)] = [row_num, url, "(error)", status,
                                   parent, depth, ts, notes]
                print(f"         -> HTTP {status}")
                time.sleep(delay)
                continue

            if "html" not in ctype.lower():
                ct = ctype.split(";")[0].strip()
                notes = f"Non-HTML content ({ct})"
                df.loc[len(df)] = [row_num, url, "(non-HTML)", status,
                                   parent, depth, ts, notes]
                print(f"         -> Skipped: {notes}")
                time.sleep(delay)
                continue

        except requests.Timeout:
            df.loc[len(df)] = [row_num, url, "(timeout)", "ERR",
                               parent, depth, ts, "Request timed out"]
            print(f"         -> TIMEOUT")
            time.sleep(delay)
            continue
        except Exception as e:
            df.loc[len(df)] = [row_num, url, "(error)", "ERR",
                               parent, depth, ts, str(e)[:120]]
            print(f"         -> ERROR: {e}")
            time.sleep(delay)
            continue

        # ── Parse ──────────────────────────────────────────────────────────────
        soup  = BeautifulSoup(resp.text, "html.parser")
        ttag  = soup.find("title")
        title = ttag.get_text(strip=True) if ttag else "(no title)"

        df.loc[len(df)] = [row_num, final, title, status,
                           parent, depth, ts, ""]
        print(f"         -> OK  '{title[:65]}'")

        # ── Discover links — dead end if none are new ──────────────────────────
        known = set(df["URL"].tolist())
        new_count = 0

        for tag_name, attr in LINK_TAGS.items():
            for tag in soup.find_all(tag_name, **{attr: True}):
                href = tag.get(attr, "").strip()
                if not href or href.startswith(("mailto:", "tel:", "javascript:", "#")):
                    continue
                abs_url = normalize(urljoin(final, href))
                if not same_domain(abs_url, base):
                    continue
                if not is_webpage(abs_url):
                    continue
                if abs_url not in known and abs_url not in queued:
                    if max_depth == -1 or (depth + 1) <= max_depth:
                        enqueue(abs_url, final, depth + 1)
                        new_count += 1

        if new_count:
            print(f"         -> +{new_count} new link(s) | queue: {len(queue)}")
        else:
            print(f"         -> Dead end (all links already recorded)")

        time.sleep(delay)

    # Clean index numbering
    df["#"] = range(1, len(df) + 1)
    return df

# ── Excel Helpers ──────────────────────────────────────────────────────────────

def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)

def _border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_header_row(ws, row: int, values: list, widths: list):
    for c_idx, (h, w) in enumerate(zip(values, widths), 1):
        cell = ws.cell(row=row, column=c_idx, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[row].height = 22

def _apply_data_row(ws, row: int, values: list, bg: str,
                    wrap_cols: set = None, red_col: int = None):
    wrap_cols = wrap_cols or {2, 3, 5, 8}
    for c_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=c_idx, value=val)
        fc   = "CC0000" if (c_idx == red_col) else "222222"
        cell.font      = Font(name="Arial", size=9, color=fc)
        cell.fill      = _fill(bg)
        cell.alignment = Alignment(vertical="top",
                                   wrap_text=(c_idx in wrap_cols))
        cell.border    = _border()
    ws.row_dimensions[row].height = 16

# ── CSV Export ─────────────────────────────────────────────────────────────────

def export_csv(df: pd.DataFrame, path: str):
    """
    Write the crawl DataFrame to a plain CSV.
    Column order matches what alt_text.py (WCAG 1.1.1 checker) expects:
      URL, Page Title, HTTP Status  (plus the remaining audit columns).
    """
    out_cols = ["URL", "Page Title", "HTTP Status",
                "Found On (Parent URL)", "Depth", "Discovered At", "Notes"]
    df[out_cols].to_csv(path, index=False, encoding="utf-8")
    print(f"\n  CSV report saved -> {path}")
    print(f"  Rows: {len(df)}  |  Ready for: python alt_text.py {path} wcag_report.csv")


# ── Excel Export ───────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, start_url: str, path: str):
    from openpyxl import Workbook

    wb    = Workbook()
    ok_df = df[df["HTTP Status"] == 200]
    er_df = df[df["HTTP Status"] != 200]
    domain = urlparse(start_url).netloc

    headers    = ["#", "URL", "Page Title", "HTTP Status",
                  "Found On (Parent URL)", "Depth", "Discovered At", "Notes"]
    col_widths = [6, 60, 42, 12, 50, 8, 14, 35]

    # ── Sheet 1 : Crawl Results ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Crawl Results"

    # Title banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "GIGW 3.0 — Web Crawler Report"
    ws["A1"].font      = Font(name="Arial", bold=True, size=16, color=C_HEADER_FG)
    ws["A1"].fill      = _fill(C_TITLE_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    # Meta row
    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Domain: {domain}   |   Start URL: {start_url}   |   "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   "
        f"Total Recorded: {len(df)}   |   OK (200): {len(ok_df)}   |   "
        f"Errors / Skipped: {len(er_df)}"
    )
    ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="444444")
    ws["A2"].fill      = _fill(C_META_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[2].height = 24

    # Header row
    _apply_header_row(ws, 3, headers, col_widths)

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), 4):
        status = row["HTTP Status"]
        is_200 = (status == 200)
        is_err = (str(status).startswith(("4", "5")) or status in ("ERR",))
        bg = C_OK_BG if is_200 else (C_ERR_BG if is_err else C_WARN_BG)
        if is_200 and (r_idx % 2 == 0):
            bg = C_ALT_BG

        row_vals = [row["#"], row["URL"], row["Page Title"], row["HTTP Status"],
                    row["Found On (Parent URL)"], row["Depth"],
                    row["Discovered At"], row["Notes"]]
        red_col = 4 if is_err else None
        _apply_data_row(ws, r_idx, row_vals, bg, red_col=red_col)

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:H{3 + len(df)}"

    # ── Sheet 2 : Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 22

    def s_banner(r, text, bg=C_TITLE_BG):
        ws2.merge_cells(f"A{r}:B{r}")
        c = ws2.cell(row=r, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
        c.fill      = _fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 28

    def s_section(r, text):
        ws2.merge_cells(f"A{r}:B{r}")
        c = ws2.cell(row=r, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[r].height = 20

    def s_kv(r, label, value, highlight=False):
        bg = C_WARN_BG if highlight else ("FFFFFF" if r % 2 else C_ALT_BG)
        a  = ws2.cell(row=r, column=1, value=label)
        b  = ws2.cell(row=r, column=2, value=value)
        for c in (a, b):
            c.font      = Font(name="Arial", size=10)
            c.fill      = _fill(bg)
            c.border    = _border()
            c.alignment = Alignment(vertical="center", indent=1)
        b.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 18

    s_banner(1, "Crawl Summary")
    s_section(2, "General Information")
    s_kv(3,  "Domain",           domain)
    s_kv(4,  "Start URL",        start_url)
    s_kv(5,  "Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    s_section(7, "Page Counts")
    s_kv(8,  "Total Entries Recorded",       len(df))
    s_kv(9,  "Successfully Loaded (HTTP 200)", len(ok_df))
    s_kv(10, "Errors / Skipped",              len(er_df),
             highlight=(len(er_df) > 0))

    row_cursor = 12
    if not ok_df.empty:
        s_section(row_cursor, "Pages by Depth Level")
        row_cursor += 1
        depth_counts = ok_df["Depth"].value_counts().sort_index()
        for d, cnt in depth_counts.items():
            s_kv(row_cursor, f"  Depth {d}", cnt)
            row_cursor += 1

    row_cursor += 1
    s_section(row_cursor, "HTTP Status Breakdown")
    row_cursor += 1
    for st, cnt in df["HTTP Status"].value_counts().items():
        hi = str(st).startswith(("4", "5")) or st == "ERR"
        s_kv(row_cursor, f"  HTTP {st}", cnt, highlight=hi)
        row_cursor += 1

    # ── Sheet 3 : Errors & Skipped ────────────────────────────────────────────
    if not er_df.empty:
        ws3 = wb.create_sheet("Errors & Skipped")

        ws3.merge_cells("A1:H1")
        ws3["A1"] = "Errors & Skipped Pages"
        ws3["A1"].font      = Font(name="Arial", bold=True, size=14,
                                   color=C_HEADER_FG)
        ws3["A1"].fill      = _fill("8B1A1A")
        ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[1].height = 28

        _apply_header_row(ws3, 2, headers, col_widths)

        for r_idx, (_, row) in enumerate(er_df.iterrows(), 3):
            row_vals = [row["#"], row["URL"], row["Page Title"],
                        row["HTTP Status"], row["Found On (Parent URL)"],
                        row["Depth"], row["Discovered At"], row["Notes"]]
            _apply_data_row(ws3, r_idx, row_vals, C_ERR_BG, red_col=4)

        ws3.freeze_panes = "A3"
        ws3.auto_filter.ref = f"A2:H{2 + len(er_df)}"

    wb.save(path)
    print(f"\n  Excel report saved -> {path}")
    print(f"  Sheets: 'Crawl Results', 'Summary'"
          + (", 'Errors & Skipped'" if not er_df.empty else ""))

# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="GIGW 3.0 Web Crawler — exhaustive internal link discovery",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python crawler.py
  python crawler.py --url https://example.gov.in
  python crawler.py --url https://example.gov.in --depth 3
  python crawler.py --url https://example.gov.in --delay 2 --output report.csv
  python crawler.py --url https://example.gov.in --format xlsx --output report.xlsx

Pipeline (crawl → WCAG alt-text check):
  python crawler.py --url https://example.gov.in --output crawl.csv
  python alt_text.py crawl.csv wcag_report.csv
        """
    )
    p.add_argument("--url",     type=str,   help="Starting URL")
    p.add_argument("--depth",   type=int,   default=-1,
                   help="Max crawl depth (-1 = unlimited, default: -1)")
    p.add_argument("--delay",   type=float, default=DEFAULT_DELAY,
                   help=f"Seconds between requests (default: {DEFAULT_DELAY})")
    p.add_argument("--timeout", type=int,   default=DEFAULT_TIMEOUT,
                   help=f"Request timeout seconds (default: {DEFAULT_TIMEOUT})")
    p.add_argument("--format",  type=str,   default="csv",
                   choices=["csv", "xlsx"],
                   help="Output format: csv (default) or xlsx")
    p.add_argument("--output",  type=str,   default=None,
                   help="Output file path (default: <domain>_crawl.csv or .xlsx)")
    return p.parse_args()

def main():
    args = parse_args()
    url  = args.url
    if not url:
        url = input("Enter the website URL to crawl: ").strip()
    if not url:
        print("[ERROR] No URL provided.")
        sys.exit(1)
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    domain  = urlparse(url).netloc.replace(":", "_")
    fmt     = args.format.lower()                          # "csv" or "xlsx"
    ext     = "." + fmt
    outfile = args.output or f"{domain}_crawl{ext}"

    # If the user gave --output without an extension, append the right one
    if args.output and "." not in args.output.rsplit("/", 1)[-1]:
        outfile = args.output + ext

    t_start = datetime.now()
    df      = crawl(url, delay=args.delay, timeout=args.timeout,
                    max_depth=args.depth)
    t_end   = datetime.now()

    ok_count  = len(df[df["HTTP Status"] == 200])
    err_count = len(df[df["HTTP Status"] != 200])
    duration  = (t_end - t_start).total_seconds()

    print(f"\n{'='*65}")
    print(f"  CRAWL COMPLETE")
    print(f"{'='*65}")
    print(f"  Total entries : {len(df)}")
    print(f"  OK (200)      : {ok_count}")
    print(f"  Errors/Skipped: {err_count}")
    print(f"  Duration      : {duration:.1f}s")
    print(f"{'='*65}")

    if fmt == "xlsx":
        export_excel(df, url, outfile)
    else:
        export_csv(df, outfile)

if __name__ == "__main__":
    main()