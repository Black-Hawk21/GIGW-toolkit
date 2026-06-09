"""
GIGW 3.0 Media Crawler
======================
Reads a crawl CSV produced by crawler.py (GIGW 3.0 Web Crawler) and
visits every successfully-loaded page (HTTP 200) to find all embedded
media items: images, videos, audio, iframes, embeds, objects, and
SVGs.

For each page the script records:
  • Parent URL  — the page being inspected
  • Media Type  — image / video / audio / iframe / embed / object / svg
  • Media URL   — absolute URL to the media file (or src/data attribute value)
  • Tag         — the HTML tag that contained the reference
  • Alt / Title — alt text (images) or title attribute where present
  • HTTP Status — HEAD-request status for the media URL (or N/A for inline SVG)

Output:
  A CSV with one row per media item plus a summary CSV with per-page counts.

Dependencies:
    pip install requests beautifulsoup4 pandas openpyxl

Usage:
    python media_crawler.py
    python media_crawler.py --input crawl.csv
    python media_crawler.py --input crawl.csv --output media_report.csv
    python media_crawler.py --input crawl.csv --format xlsx
    python media_crawler.py --input crawl.csv --delay 0.5 --timeout 10
    python media_crawler.py --input crawl.csv --no-verify
"""

import argparse
import sys
import time
from urllib.parse import urljoin, urlparse, urldefrag
from datetime import datetime

import requests
import pandas as pd

try:
    from bs4 import BeautifulSoup
except ImportError:
    print("[ERROR] Run: pip install requests beautifulsoup4 pandas openpyxl")
    sys.exit(1)

# ── Constants ──────────────────────────────────────────────────────────────────

DEFAULT_DELAY      = 0.5
DEFAULT_TIMEOUT    = 10
DEFAULT_USER_AGENT = "GIGW-MediaCrawler/1.0 (Government Website Evaluation Tool)"

# Tags and their src-like attributes that point to media resources
MEDIA_TAG_ATTRS = {
    "img":    {"src": "image",   "srcset": "image"},
    "video":  {"src": "video",   "poster": "image"},
    "source": {"src": "video"},   # <source> inside <video> or <audio>
    "audio":  {"src": "audio"},
    "track":  {"src": "audio"},   # captions / subtitles
    "iframe": {"src": "iframe"},
    "embed":  {"src": "embed"},
    "object": {"data": "object"},
}

# ── Colour palette (Excel) ─────────────────────────────────────────────────────
C_TITLE_BG  = "0D2137"
C_HEADER_BG = "1A3C5E"
C_HEADER_FG = "FFFFFF"
C_ALT_BG    = "F7FAFD"
C_META_BG   = "F0F4F8"
C_BORDER    = "C5D5E4"
C_OK_BG     = "EAF4EC"
C_ERR_BG    = "FCE8E8"
C_WARN_BG   = "FFF8E1"

# ── Helpers ────────────────────────────────────────────────────────────────────

def normalize(url: str) -> str:
    url, _ = urldefrag(url)
    return url.rstrip("/")


def is_inline_svg(tag) -> bool:
    """Return True for an <svg> element that is embedded directly in the HTML."""
    return tag.name == "svg"


def extract_srcset_urls(srcset_val: str, base_url: str) -> list[str]:
    """Parse a srcset attribute and return a list of absolute URLs."""
    urls = []
    for part in srcset_val.split(","):
        part = part.strip()
        if not part:
            continue
        candidate = part.split()[0]  # first token is the URL
        abs_url = urljoin(base_url, candidate)
        urls.append(normalize(abs_url))
    return urls


def detect_media_type_from_parent(tag) -> str:
    """
    A <source> tag inherits its media type from its parent container:
    <video> → video, <audio> → audio, anything else → embed.
    """
    parent = tag.parent
    if parent:
        if parent.name == "video":
            return "video"
        if parent.name == "audio":
            return "audio"
    return "embed"


def head_status(session: requests.Session, url: str, timeout: int) -> str:
    """Return the HTTP status code (as str) for a HEAD request, or an error label."""
    try:
        r = session.head(url, timeout=timeout, allow_redirects=True)
        return str(r.status_code)
    except requests.Timeout:
        return "TIMEOUT"
    except Exception:
        return "ERR"


# ── Core scanner ──────────────────────────────────────────────────────────────

def scan_page(
    page_url: str,
    session: requests.Session,
    timeout: int,
    verify_media: bool,
) -> list[dict]:
    """
    Fetch one page and return a list of media-item dicts found on it.
    Returns an empty list on fetch failure.
    """
    try:
        resp = session.get(page_url, timeout=timeout, allow_redirects=True)
        if resp.status_code != 200:
            return []
        if "html" not in resp.headers.get("Content-Type", "").lower():
            return []
    except Exception:
        return []

    soup  = BeautifulSoup(resp.text, "html.parser")
    items = []

    # ── 1. Standard tag-based media ───────────────────────────────────────────
    for tag_name, attr_map in MEDIA_TAG_ATTRS.items():
        for tag in soup.find_all(tag_name):

            # Resolve media type for <source> based on parent context
            for attr, media_type in attr_map.items():
                val = tag.get(attr, "").strip()
                if not val:
                    continue

                if tag_name == "source":
                    media_type = detect_media_type_from_parent(tag)

                # srcset may contain multiple URLs
                if attr == "srcset":
                    for abs_url in extract_srcset_urls(val, page_url):
                        items.append({
                            "parent_url":   page_url,
                            "media_type":   media_type,
                            "media_url":    abs_url,
                            "tag":          f"<{tag_name} srcset>",
                            "alt_title":    tag.get("alt", tag.get("title", "")).strip(),
                            "http_status":  head_status(session, abs_url, timeout)
                                            if verify_media else "N/A",
                        })
                else:
                    abs_url = normalize(urljoin(page_url, val))
                    items.append({
                        "parent_url":   page_url,
                        "media_type":   media_type,
                        "media_url":    abs_url,
                        "tag":          f"<{tag_name}>",
                        "alt_title":    tag.get("alt", tag.get("title", "")).strip(),
                        "http_status":  head_status(session, abs_url, timeout)
                                        if verify_media else "N/A",
                    })

    # ── 2. Inline SVG elements ────────────────────────────────────────────────
    for svg in soup.find_all("svg"):
        items.append({
            "parent_url":  page_url,
            "media_type":  "svg (inline)",
            "media_url":   "(inline — no external URL)",
            "tag":         "<svg>",
            "alt_title":   svg.get("aria-label", svg.get("title", "")).strip(),
            "http_status": "N/A",
        })

    # ── 3. CSS background images referenced via <style> blocks ───────────────
    import re
    style_blocks = soup.find_all("style")
    for style in style_blocks:
        for match in re.finditer(r'url\(["\']?([^"\')\s]+)["\']?\)', style.get_text()):
            raw = match.group(1).strip()
            if raw.startswith("data:"):
                continue
            abs_url = normalize(urljoin(page_url, raw))
            items.append({
                "parent_url":  page_url,
                "media_type":  "image (CSS bg)",
                "media_url":   abs_url,
                "tag":         "<style> url()",
                "alt_title":   "",
                "http_status": head_status(session, abs_url, timeout)
                               if verify_media else "N/A",
            })

    # ── 4. Inline style attributes with background-image ─────────────────────
    for tag in soup.find_all(style=True):
        for match in re.finditer(
            r'background(?:-image)?\s*:\s*url\(["\']?([^"\')\s]+)["\']?\)',
            tag["style"],
        ):
            raw = match.group(1).strip()
            if raw.startswith("data:"):
                continue
            abs_url = normalize(urljoin(page_url, raw))
            items.append({
                "parent_url":  page_url,
                "media_type":  "image (CSS bg)",
                "media_url":   abs_url,
                "tag":         f"<{tag.name} style>",
                "alt_title":   tag.get("alt", tag.get("title", "")).strip(),
                "http_status": head_status(session, abs_url, timeout)
                               if verify_media else "N/A",
            })

    return items


# ── Main crawler loop ─────────────────────────────────────────────────────────

def crawl_media(
    input_csv:    str,
    delay:        float,
    timeout:      int,
    verify_media: bool,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Read page URLs from the crawl CSV, scan each page for media,
    and return (detail_df, summary_df).
    """
    # Load crawl CSV
    try:
        crawl_df = pd.read_csv(input_csv)
    except FileNotFoundError:
        print(f"[ERROR] Input CSV not found: {input_csv}")
        sys.exit(1)

    if "URL" not in crawl_df.columns:
        print("[ERROR] Input CSV must have a 'URL' column (output of crawler.py).")
        sys.exit(1)

    # Keep only successfully crawled pages
    if "HTTP Status" in crawl_df.columns:
        pages_df = crawl_df[crawl_df["HTTP Status"] == 200].copy()
        skipped  = len(crawl_df) - len(pages_df)
    else:
        pages_df = crawl_df.copy()
        skipped  = 0

    urls = pages_df["URL"].dropna().unique().tolist()

    session = requests.Session()
    session.headers.update({
        "User-Agent": DEFAULT_USER_AGENT,
        "Accept":     "text/html,*/*;q=0.8",
    })

    all_items: list[dict] = []

    print(f"\n{'='*65}")
    print(f"  GIGW 3.0 Media Crawler")
    print(f"{'='*65}")
    print(f"  Input CSV     : {input_csv}")
    print(f"  Pages to scan : {len(urls)}  (skipped {skipped} non-200 entries)")
    print(f"  Verify media  : {'Yes (HEAD requests)' if verify_media else 'No'}")
    print(f"  Delay         : {delay}s")
    print(f"{'='*65}\n")

    for idx, url in enumerate(urls, 1):
        print(f"  [{idx:>4}/{len(urls)}] {url}")
        items = scan_page(url, session, timeout, verify_media)
        all_items.extend(items)

        counts_by_type = {}
        for item in items:
            counts_by_type[item["media_type"]] = (
                counts_by_type.get(item["media_type"], 0) + 1
            )
        summary_str = "  ".join(f"{t}: {c}" for t, c in counts_by_type.items())
        print(f"         -> {len(items)} media item(s)  |  {summary_str or 'none'}")

        time.sleep(delay)

    # ── Build detail DataFrame ─────────────────────────────────────────────────
    detail_df = pd.DataFrame(all_items, columns=[
        "parent_url", "media_type", "media_url",
        "tag", "alt_title", "http_status",
    ])
    detail_df.index = range(1, len(detail_df) + 1)
    detail_df.index.name = "#"

    # ── Build summary DataFrame ────────────────────────────────────────────────
    if not detail_df.empty:
        summary_rows = []
        for page_url in urls:
            page_items = detail_df[detail_df["parent_url"] == page_url]
            total      = len(page_items)
            by_type    = page_items["media_type"].value_counts().to_dict()
            summary_rows.append({
                "Page URL":       page_url,
                "Total Media":    total,
                "Images":         by_type.get("image", 0)
                                  + by_type.get("image (CSS bg)", 0),
                "Videos":         by_type.get("video", 0),
                "Audio":          by_type.get("audio", 0),
                "Iframes":        by_type.get("iframe", 0),
                "Embeds/Objects": by_type.get("embed", 0) + by_type.get("object", 0),
                "SVG (inline)":   by_type.get("svg (inline)", 0),
            })
    else:
        summary_rows = [
            {"Page URL": u, "Total Media": 0, "Images": 0, "Videos": 0,
             "Audio": 0, "Iframes": 0, "Embeds/Objects": 0, "SVG (inline)": 0}
            for u in urls
        ]

    summary_df = pd.DataFrame(summary_rows)

    return detail_df, summary_df


# ── CSV Export ─────────────────────────────────────────────────────────────────

def export_csv(detail_df: pd.DataFrame, summary_df: pd.DataFrame,
               detail_path: str, summary_path: str):
    detail_df.to_csv(detail_path, index=True, encoding="utf-8")
    summary_df.to_csv(summary_path, index=False, encoding="utf-8")
    print(f"\n  Detail CSV  -> {detail_path}  ({len(detail_df)} rows)")
    print(f"  Summary CSV -> {summary_path}  ({len(summary_df)} pages)")


# ── Excel Export ───────────────────────────────────────────────────────────────

def export_excel(detail_df: pd.DataFrame, summary_df: pd.DataFrame,
                 input_csv: str, path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def _fill(hex_colour):
        return PatternFill("solid", fgColor=hex_colour)

    def _border():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    def _header_row(ws, row, values, widths):
        for c, (h, w) in enumerate(zip(values, widths), 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
            cell.fill      = _fill(C_HEADER_BG)
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border    = _border()
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[row].height = 22

    def _data_row(ws, row, values, bg, wrap_cols=None):
        wrap_cols = wrap_cols or set()
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = _fill(bg)
            cell.alignment = Alignment(vertical="top",
                                       wrap_text=(c in wrap_cols))
            cell.border    = _border()
        ws.row_dimensions[row].height = 16

    wb = Workbook()

    # ── Sheet 1 : Media Detail ────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Media Detail"

    ws1.merge_cells("A1:F1")
    ws1["A1"] = "GIGW 3.0 — Media Crawler Report"
    ws1["A1"].font      = Font(name="Arial", bold=True, size=16, color=C_HEADER_FG)
    ws1["A1"].fill      = _fill(C_TITLE_BG)
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 34

    ws1.merge_cells("A2:F2")
    ws1["A2"] = (
        f"Source CSV: {input_csv}   |   "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   "
        f"Total media items: {len(detail_df)}"
    )
    ws1["A2"].font      = Font(name="Arial", size=9, italic=True, color="444444")
    ws1["A2"].fill      = _fill(C_META_BG)
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 20

    headers1    = ["#", "Parent URL", "Media Type", "Media URL",
                   "Tag", "Alt / Title", "HTTP Status"]
    col_widths1 = [6,   60,            16,           60, 18, 30, 12]
    _header_row(ws1, 3, headers1, col_widths1)

    TYPE_COLOURS = {
        "image":           C_OK_BG,
        "image (CSS bg)":  "F0FAF0",
        "video":           "EEF0FA",
        "audio":           "FAF0EE",
        "iframe":          C_WARN_BG,
        "embed":           "FFF0E1",
        "object":          "FFF0E1",
        "svg (inline)":    "F5F0FA",
    }

    for r_idx, (row_num, row) in enumerate(detail_df.iterrows(), 4):
        bg = TYPE_COLOURS.get(row["media_type"], C_ALT_BG)
        status = str(row["http_status"])
        if status.startswith(("4", "5")) or status in ("ERR", "TIMEOUT"):
            bg = C_ERR_BG
        vals = [
            row_num,
            row["parent_url"],
            row["media_type"],
            row["media_url"],
            row["tag"],
            row["alt_title"],
            row["http_status"],
        ]
        _data_row(ws1, r_idx, vals, bg, wrap_cols={2, 4, 6})

    ws1.freeze_panes = "A4"
    ws1.auto_filter.ref = f"A3:G{3 + len(detail_df)}"

    # ── Sheet 2 : Page Summary ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Page Summary")

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "Media Count per Page"
    ws2["A1"].font      = Font(name="Arial", bold=True, size=14, color=C_HEADER_FG)
    ws2["A1"].fill      = _fill(C_TITLE_BG)
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    headers2    = ["#", "Page URL", "Total Media", "Images",
                   "Videos", "Audio", "Iframes", "Embeds/Objects", "SVG (inline)"]
    col_widths2 = [6,   65,          14,            10, 10, 10, 10, 16, 14]
    _header_row(ws2, 2, headers2, col_widths2)

    for r_idx, (_, row) in enumerate(summary_df.iterrows(), 3):
        total = row["Total Media"]
        bg    = C_OK_BG if total > 0 else C_ALT_BG
        if r_idx % 2 == 0 and total > 0:
            bg = "D8EFD8"
        vals = [
            r_idx - 2,
            row["Page URL"],
            row["Total Media"],
            row["Images"],
            row["Videos"],
            row["Audio"],
            row["Iframes"],
            row["Embeds/Objects"],
            row["SVG (inline)"],
        ]
        _data_row(ws2, r_idx, vals, bg, wrap_cols={2})

    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:I{2 + len(summary_df)}"

    # ── Sheet 3 : Overall Stats ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Stats")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 20

    def stat_banner(r, text):
        ws3.merge_cells(f"A{r}:B{r}")
        c = ws3.cell(row=r, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=13, color=C_HEADER_FG)
        c.fill      = _fill(C_TITLE_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[r].height = 26

    def stat_section(r, text):
        ws3.merge_cells(f"A{r}:B{r}")
        c = ws3.cell(row=r, column=1, value=text)
        c.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
        c.fill      = _fill(C_HEADER_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws3.row_dimensions[r].height = 20

    def stat_kv(r, label, value):
        bg = "FFFFFF" if r % 2 else C_ALT_BG
        a  = ws3.cell(row=r, column=1, value=label)
        b  = ws3.cell(row=r, column=2, value=value)
        for cell in (a, b):
            cell.font      = Font(name="Arial", size=10)
            cell.fill      = _fill(bg)
            cell.border    = _border()
            cell.alignment = Alignment(vertical="center", indent=1)
        b.alignment = Alignment(horizontal="center", vertical="center")
        ws3.row_dimensions[r].height = 18

    stat_banner(1, "Media Scan Statistics")
    stat_section(2, "Source")
    stat_kv(3, "Input CSV",       input_csv)
    stat_kv(4, "Pages Scanned",   len(summary_df))
    stat_kv(5, "Generated",       datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    stat_section(7, "Media Totals")
    stat_kv(8,  "Total Media Items",   len(detail_df))
    stat_kv(9,  "Images",              int(summary_df["Images"].sum()))
    stat_kv(10, "Videos",              int(summary_df["Videos"].sum()))
    stat_kv(11, "Audio",               int(summary_df["Audio"].sum()))
    stat_kv(12, "Iframes",             int(summary_df["Iframes"].sum()))
    stat_kv(13, "Embeds / Objects",    int(summary_df["Embeds/Objects"].sum()))
    stat_kv(14, "SVG (inline)",        int(summary_df["SVG (inline)"].sum()))

    if not detail_df.empty:
        stat_section(16, "Pages with Most Media")
        top5 = summary_df.nlargest(5, "Total Media")[["Page URL", "Total Media"]]
        for i, (_, row) in enumerate(top5.iterrows(), 17):
            stat_kv(i, row["Page URL"][:55], row["Total Media"])

    wb.save(path)
    print(f"\n  Excel report saved -> {path}")
    print(f"  Sheets: 'Media Detail', 'Page Summary', 'Stats'")


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="GIGW 3.0 Media Crawler — scan pages for embedded media",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python media_crawler.py
  python media_crawler.py --input crawl.csv
  python media_crawler.py --input crawl.csv --output media_report.csv
  python media_crawler.py --input crawl.csv --format xlsx
  python media_crawler.py --input crawl.csv --delay 1 --timeout 15
  python media_crawler.py --input crawl.csv --no-verify

Pipeline:
  python crawler.py --url https://example.gov.in --output crawl.csv
  python media_crawler.py --input crawl.csv
        """,
    )
    p.add_argument("--input",     type=str,  default=None,
                   help="Path to crawl CSV from crawler.py (default: prompt)")
    p.add_argument("--output",    type=str,  default=None,
                   help="Base output path (extensions added automatically)")
    p.add_argument("--format",    type=str,  default="csv",
                   choices=["csv", "xlsx"],
                   help="Output format: csv (default) or xlsx")
    p.add_argument("--delay",     type=float, default=DEFAULT_DELAY,
                   help=f"Seconds between page requests (default: {DEFAULT_DELAY})")
    p.add_argument("--timeout",   type=int,   default=DEFAULT_TIMEOUT,
                   help=f"Request timeout in seconds (default: {DEFAULT_TIMEOUT})")
    p.add_argument("--no-verify", dest="verify", action="store_false",
                   help="Skip HEAD requests for media URLs (faster, no status codes)")
    p.set_defaults(verify=True)
    return p.parse_args()


def main():
    args = parse_args()

    input_csv = args.input
    if not input_csv:
        input_csv = input("Enter path to the crawl CSV: ").strip()
    if not input_csv:
        print("[ERROR] No input CSV provided.")
        sys.exit(1)

    fmt = args.format.lower()
    ext = "." + fmt

    # Derive output base name from input CSV if not provided
    if args.output:
        base = args.output
        if "." in base.rsplit("/", 1)[-1]:
            base = base.rsplit(".", 1)[0]   # strip any extension the user typed
    else:
        base = input_csv.replace(".csv", "").replace(".xlsx", "") + "_media"

    t_start = datetime.now()
    detail_df, summary_df = crawl_media(
        input_csv    = input_csv,
        delay        = args.delay,
        timeout      = args.timeout,
        verify_media = args.verify,
    )
    t_end = datetime.now()

    duration  = (t_end - t_start).total_seconds()
    total     = len(detail_df)
    pages     = len(summary_df)
    with_media = int((summary_df["Total Media"] > 0).sum()) if not summary_df.empty else 0

    print(f"\n{'='*65}")
    print(f"  SCAN COMPLETE")
    print(f"{'='*65}")
    print(f"  Pages scanned     : {pages}")
    print(f"  Pages with media  : {with_media}")
    print(f"  Total media items : {total}")
    if not detail_df.empty:
        for mtype, cnt in detail_df["media_type"].value_counts().items():
            print(f"    {mtype:<22}: {cnt}")
    print(f"  Duration          : {duration:.1f}s")
    print(f"{'='*65}")

    if fmt == "xlsx":
        out_path = base + ext
        export_excel(detail_df, summary_df, input_csv, out_path)
    else:
        detail_path  = base + "_detail.csv"
        summary_path = base + "_summary.csv"
        export_csv(detail_df, summary_df, detail_path, summary_path)


if __name__ == "__main__":
    main()
