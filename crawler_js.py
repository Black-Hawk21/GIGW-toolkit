"""
GIGW 3.0 Web Crawler (JS-capable)
==================================
Same purpose as crawler.py but uses Playwright (headless Chromium) to
fully render JavaScript before extracting links.  This means it can
crawl single-page apps (React, Angular, Vue), sites that load
navigation via AJAX, and any page that injects <a> tags at runtime.

The output CSV is **identical** to crawler.py so it feeds directly
into alt_text.py, contrast_checker.py, and media_crawler.py.

Dependencies:
    pip install playwright pandas openpyxl
    python -m playwright install chromium

Usage:
    python crawler_js.py
    python crawler_js.py --url https://example.gov.in
    python crawler_js.py --url https://example.gov.in --depth 3
    python crawler_js.py --url https://example.gov.in --delay 2 --output crawl.csv

Pipeline (crawl -> WCAG check):
    python crawler_js.py --url https://example.gov.in --output crawl.csv
    python alt_text.py crawl.csv wcag_report.csv
"""

import argparse
import sys
import time
from urllib.parse import urljoin, urlparse, urldefrag
from collections import deque
from datetime import datetime

import pandas as pd

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PwTimeout
except ImportError:
    print("[ERROR] Run:  pip install playwright && python -m playwright install chromium")
    sys.exit(1)

# ── Constants ──────────────────────────────────────────────────────────────────

DEFAULT_DELAY       = 1.5
DEFAULT_TIMEOUT     = 20       # seconds — page navigation timeout
DEFAULT_WAIT_AFTER  = 2000     # ms — extra wait for JS after load event
DEFAULT_USER_AGENT  = "GIGW-Crawler/3.0-JS (Government Website Evaluation Tool)"

SKIP_EXTENSIONS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".zip", ".rar", ".tar", ".gz",
    ".jpg", ".jpeg", ".png", ".gif", ".bmp", ".svg", ".webp", ".ico",
    ".mp3", ".mp4", ".avi", ".mov", ".wmv",
    ".css", ".js", ".json", ".xml",
    ".woff", ".woff2", ".ttf", ".eot",
}

# ── URL Utilities ──────────────────────────────────────────────────────────────

def normalize(url: str) -> str:
    url, _ = urldefrag(url)
    return url.rstrip("/")


def same_domain(url: str, base: str) -> bool:
    netloc = urlparse(url).netloc.lower()
    return netloc == base or netloc.endswith("." + base)


def is_webpage(url: str) -> bool:
    path = urlparse(url).path.lower()
    last = path.rsplit("/", 1)[-1]
    ext  = "." + last.rsplit(".", 1)[-1] if "." in last else ""
    return ext not in SKIP_EXTENSIONS


# ── Link Extraction (runs inside the browser page) ────────────────────────────

EXTRACT_LINKS_JS = """
() => {
    const results = new Set();
    // Standard <a>, <link>, <area> tags
    document.querySelectorAll('a[href], link[href], area[href]').forEach(el => {
        const href = el.getAttribute('href');
        if (href) results.add(href);
    });
    // Also check onclick/data-href patterns used by some SPAs
    document.querySelectorAll('[data-href]').forEach(el => {
        const href = el.getAttribute('data-href');
        if (href) results.add(href);
    });
    return [...results];
}
"""

EXTRACT_TITLE_JS = """
() => document.title || '(no title)'
"""


# ── Crawler ────────────────────────────────────────────────────────────────────

def crawl(start_url: str, delay: float = DEFAULT_DELAY,
          timeout: int = DEFAULT_TIMEOUT,
          wait_after: int = DEFAULT_WAIT_AFTER,
          max_depth: int = -1,
          headless: bool = True) -> pd.DataFrame:

    base = urlparse(start_url).netloc.lower()

    columns = ["#", "URL", "Page Title", "HTTP Status",
               "Found On (Parent URL)", "Depth", "Discovered At", "Notes"]
    df = pd.DataFrame(columns=columns)

    queue  = deque()       # (url, parent_url, depth)
    queued = set()

    def enqueue(url, parent, depth):
        nurl = normalize(url)
        if nurl not in queued and nurl not in df["URL"].values:
            queued.add(nurl)
            queue.append((nurl, parent, depth))

    enqueue(normalize(start_url), "---", 0)

    depth_label = "Unlimited (exhaustive)" if max_depth == -1 else str(max_depth)
    print(f"\n{'='*65}")
    print(f"  GIGW 3.0 Web Crawler (JS-capable)")
    print(f"{'='*65}")
    print(f"  Start URL  : {start_url}")
    print(f"  Domain     : {base}")
    print(f"  Max Depth  : {depth_label}")
    print(f"  JS Wait    : {wait_after} ms")
    print(f"  Headless   : {headless}")
    print(f"{'='*65}\n")

    interrupted = False
    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=headless)
        context = browser.new_context(
            user_agent=DEFAULT_USER_AGENT,
            viewport={"width": 1280, "height": 800},
            ignore_https_errors=True,
        )
        page = context.new_page()

        try:
            while queue:
                url, parent, depth = queue.popleft()

                if url in df["URL"].values:
                    continue

                row_num = len(df) + 1
                ts      = datetime.now().strftime("%H:%M:%S")
                print(f"  [{row_num:>4}] (depth={depth}) {url}")

                # ── Navigate ──────────────────────────────────────────────
                try:
                    resp = page.goto(url, timeout=timeout * 1000,
                                     wait_until="domcontentloaded")

                    status = resp.status if resp else 0
                    final  = normalize(page.url)
                    notes  = ""

                    # Check for external redirect
                    if not same_domain(final, base):
                        notes = "Redirected to external domain"
                        df.loc[len(df)] = [row_num, url, "(external redirect)",
                                           status, parent, depth, ts, notes]
                        print(f"         -> Redirected externally to {final}")
                        time.sleep(delay)
                        continue

                    if status >= 400:
                        notes = f"HTTP {status}"
                        df.loc[len(df)] = [row_num, url, "(error)", status,
                                           parent, depth, ts, notes]
                        print(f"         -> HTTP {status}")
                        time.sleep(delay)
                        continue

                    # Wait for JS to finish rendering
                    try:
                        page.wait_for_load_state("networkidle", timeout=wait_after + 3000)
                    except PwTimeout:
                        pass   # networkidle is best-effort
                    page.wait_for_timeout(wait_after)

                except PwTimeout:
                    df.loc[len(df)] = [row_num, url, "(timeout)", "ERR",
                                       parent, depth, ts, "Navigation timed out"]
                    print(f"         -> TIMEOUT")
                    time.sleep(delay)
                    continue
                except Exception as e:
                    df.loc[len(df)] = [row_num, url, "(error)", "ERR",
                                       parent, depth, ts, str(e)[:120]]
                    print(f"         -> ERROR: {e}")
                    time.sleep(delay)
                    continue

                # ── Extract title ─────────────────────────────────────────
                try:
                    title = page.evaluate(EXTRACT_TITLE_JS)
                except Exception:
                    title = "(no title)"

                # Check if the final URL differs (JS-side redirect)
                final = normalize(page.url)
                if not same_domain(final, base):
                    df.loc[len(df)] = [row_num, url, title,
                                       status, parent, depth, ts,
                                       "JS-redirected to external domain"]
                    print(f"         -> JS-redirected externally")
                    time.sleep(delay)
                    continue

                df.loc[len(df)] = [row_num, final, title, status,
                                   parent, depth, ts, ""]
                print(f"         -> OK  '{title[:65]}'")

                # ── Extract links ─────────────────────────────────────────
                try:
                    raw_links = page.evaluate(EXTRACT_LINKS_JS)
                except Exception:
                    raw_links = []

                known     = set(df["URL"].tolist())
                new_count = 0

                for href in raw_links:
                    if not href or href.startswith(("mailto:", "tel:", "javascript:", "#")):
                        continue
                    abs_url = normalize(urljoin(final, href))
                    if not same_domain(abs_url, base):
                        continue
                    if not is_webpage(abs_url):
                        continue
                    if abs_url not in known and abs_url not in queued:
                        if max_depth == -1 or (depth + 1) <= max_depth:
                            enqueue(abs_url, final, depth + 1)
                            new_count += 1

                if new_count:
                    print(f"         -> +{new_count} new link(s) | queue: {len(queue)}")
                else:
                    print(f"         -> Dead end (all links already recorded)")

                time.sleep(delay)

        except KeyboardInterrupt:
            interrupted = True
            print(f"\n\n  ** Ctrl+C detected — stopping crawl **")
            print(f"  ** {len(df)} page(s) collected so far, saving... **")

        finally:
            try:
                browser.close()
            except Exception:
                pass  # browser may already be dead after Ctrl+C

    if len(df) > 0:
        df["#"] = range(1, len(df) + 1)
    return df, interrupted


# ── CSV Export ─────────────────────────────────────────────────────────────────

def export_csv(df: pd.DataFrame, path: str):
    out_cols = ["URL", "Page Title", "HTTP Status",
                "Found On (Parent URL)", "Depth", "Discovered At", "Notes"]
    df[out_cols].to_csv(path, index=False, encoding="utf-8")
    print(f"\n  CSV report saved -> {path}")
    print(f"  Rows: {len(df)}  |  Ready for: python alt_text.py {path} wcag_report.csv")


# ── Excel Export ───────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, start_url: str, path: str):
    """Identical formatting to crawler.py Excel output."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    C_TITLE_BG  = "0D2137"
    C_HEADER_BG = "1A3C5E"
    C_HEADER_FG = "FFFFFF"
    C_OK_BG     = "EAF4EC"
    C_ALT_BG    = "F7FAFD"
    C_ERR_BG    = "FCE8E8"
    C_WARN_BG   = "FFF8E1"
    C_META_BG   = "F0F4F8"
    C_BORDER    = "C5D5E4"

    def _fill(c): return PatternFill("solid", fgColor=c)
    def _bdr():
        s = Side(style="thin", color=C_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)

    wb     = Workbook()
    ok_df  = df[df["HTTP Status"] == 200]
    er_df  = df[df["HTTP Status"] != 200]
    domain = urlparse(start_url).netloc

    headers    = ["#", "URL", "Page Title", "HTTP Status",
                  "Found On (Parent URL)", "Depth", "Discovered At", "Notes"]
    col_widths = [6, 60, 42, 12, 50, 8, 14, 35]

    ws = wb.active
    ws.title = "Crawl Results"

    ws.merge_cells("A1:H1")
    ws["A1"] = "GIGW 3.0 -- Web Crawler Report (JS)"
    ws["A1"].font      = Font(name="Arial", bold=True, size=16, color=C_HEADER_FG)
    ws["A1"].fill      = _fill(C_TITLE_BG)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Domain: {domain}   |   Start URL: {start_url}   |   "
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   "
        f"Total Recorded: {len(df)}   |   OK (200): {len(ok_df)}   |   "
        f"Errors / Skipped: {len(er_df)}"
    )
    ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="444444")
    ws["A2"].fill      = _fill(C_META_BG)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 24

    # Header row
    for c_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=C_HEADER_FG)
        cell.fill      = _fill(C_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _bdr()
        ws.column_dimensions[get_column_letter(c_idx)].width = w
    ws.row_dimensions[3].height = 22

    # Data rows
    for r_idx, (_, row) in enumerate(df.iterrows(), 4):
        status = row["HTTP Status"]
        is_200 = (status == 200)
        is_err = (str(status).startswith(("4", "5")) or status in ("ERR",))
        bg = C_OK_BG if is_200 else (C_ERR_BG if is_err else C_WARN_BG)
        if is_200 and (r_idx % 2 == 0):
            bg = C_ALT_BG
        vals = [row["#"], row["URL"], row["Page Title"], row["HTTP Status"],
                row["Found On (Parent URL)"], row["Depth"],
                row["Discovered At"], row["Notes"]]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font   = Font(name="Arial", size=9)
            cell.fill   = _fill(bg)
            cell.border = _bdr()

    ws.freeze_panes = "A4"
    wb.save(path)
    print(f"\n  Excel report saved -> {path}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="GIGW 3.0 Web Crawler (JS-capable) -- headless Chromium link discovery",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python crawler_js.py
  python crawler_js.py --url https://example.gov.in
  python crawler_js.py --url https://example.gov.in --depth 3
  python crawler_js.py --url https://example.gov.in --delay 2 --output report.csv
  python crawler_js.py --url https://example.gov.in --no-headless

Pipeline (crawl -> WCAG alt-text check):
  python crawler_js.py --url https://example.gov.in --output crawl.csv
  python alt_text.py crawl.csv wcag_report.csv
        """,
    )
    p.add_argument("--url",     type=str,   help="Starting URL")
    p.add_argument("--depth",   type=int,   default=-1,
                   help="Max crawl depth (-1 = unlimited, default: -1)")
    p.add_argument("--delay",   type=float, default=DEFAULT_DELAY,
                   help=f"Seconds between page loads (default: {DEFAULT_DELAY})")
    p.add_argument("--timeout", type=int,   default=DEFAULT_TIMEOUT,
                   help=f"Navigation timeout in seconds (default: {DEFAULT_TIMEOUT})")
    p.add_argument("--wait",    type=int,   default=DEFAULT_WAIT_AFTER,
                   help=f"Extra ms to wait for JS after load (default: {DEFAULT_WAIT_AFTER})")
    p.add_argument("--format",  type=str,   default="csv",
                   choices=["csv", "xlsx"],
                   help="Output format: csv (default) or xlsx")
    p.add_argument("--output",  type=str,   default=None,
                   help="Output file path (default: <domain>_crawl_js.csv)")
    p.add_argument("--no-headless", dest="headless", action="store_false",
                   help="Show the browser window (useful for debugging)")
    p.set_defaults(headless=True)
    return p.parse_args()


def main():
    args = parse_args()
    url  = args.url
    if not url:
        url = input("Enter the website URL to crawl: ").strip()
    if not url:
        print("[ERROR] No URL provided.")
        sys.exit(1)
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    domain  = urlparse(url).netloc.replace(":", "_")
    fmt     = args.format.lower()
    ext     = "." + fmt
    outfile = args.output or f"{domain}_crawl_js{ext}"

    if args.output and "." not in args.output.rsplit("/", 1)[-1]:
        outfile = args.output + ext

    t_start = datetime.now()
    df, interrupted = crawl(url, delay=args.delay, timeout=args.timeout,
                            wait_after=args.wait, max_depth=args.depth,
                            headless=args.headless)
    t_end = datetime.now()

    ok_count  = len(df[df["HTTP Status"] == 200]) if len(df) > 0 else 0
    err_count = len(df[df["HTTP Status"] != 200]) if len(df) > 0 else 0
    duration  = (t_end - t_start).total_seconds()

    label = "CRAWL INTERRUPTED (partial save)" if interrupted else "CRAWL COMPLETE"

    print(f"\n{'='*65}")
    print(f"  {label}")
    print(f"{'='*65}")
    print(f"  Total entries : {len(df)}")
    print(f"  OK (200)      : {ok_count}")
    print(f"  Errors/Skipped: {err_count}")
    print(f"  Duration      : {duration:.1f}s")
    print(f"{'='*65}")

    if len(df) == 0:
        print("\n  No data collected — nothing to save.")
        sys.exit(1)

    if fmt == "xlsx":
        export_excel(df, url, outfile)
    else:
        export_csv(df, outfile)

    if interrupted:
        sys.exit(2)   # non-zero but distinct from error (exit 1)


if __name__ == "__main__":
    main()
